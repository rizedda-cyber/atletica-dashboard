"""
data_loader.py — Modulo per il caricamento e la pulizia dei dati di corsa e VBT.

Contiene:
  - load_running_data():  parsing di "Lavori Corsa.xlsx"
  - load_vbt_data():      parsing di "VBT2026322.xlsx" (via XML diretto)
  - normalize_names():    unificazione dei nomi atleta tra i due dataset
  - parse_time():         conversione dei tempi dal formato Excel a float

Per aggiungere nuovi atleti al mapping, è sufficiente estendere il
dizionario NOME_MAPPING qui sotto.
"""

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import openpyxl

# ──────────────────────────────────────────────────────────────────────
# MAPPATURA NOMI — estendere questo dizionario per unire nomi diversi
# Formato: 'nome nel file': 'nome unificato'
# ──────────────────────────────────────────────────────────────────────
NOME_MAPPING = {
    # Corsa (Nome + Cognome già uniti) → nome unificato
    "Sisssi Corrias": "Silvia Corrias",
    "Sisssi": "Silvia Corrias",
    "Sissi": "Silvia Corrias",
    "Silvia Corrias": "Silvia Corrias",
    "Riki Murru": "Riccardo Murru",
    "Riki": "Riccardo Murru",
    "Ale Demy": "Alessandro Demicheli",
    "Manu Corda": "Manuel Corda",
    "Enrico Deidda": "Enrico Deidda",  # normalizzazione case

    # VBT (solo Nome) → nome unificato
    "Zedda": "Riccardo Zedda",
    "Sindaco": "Riccardo Zedda",
    "Murru": "Riccardo Murru",
    "Demy": "Alessandro Demicheli",
    "Manu": "Manuel Corda",
    "Alberto": "Alberto Pinna",
    "Simone": "Simone Fantuzzo",
    "Davide": "Davide Pinna",
    "Giulia": "Giulia Mannu",
    "Giulia Pati": "Giulia Pati",
    "Elena": "Elena Moccia",
    "Enrico": "Enrico Deidda",
    "Leonardo": "Leonardo Carboni",
    "Manuel": "Manuel Corda",
    "Martina": "Martina Italia",
    "Sara": "Sara Cesaraccio",
    "Giuseppe": "Giuseppe Ledda",
    "Priscilla": "Priscilla Casu",
    "Marianna": "Marianna Deidda",
    "Laura": "Laura Murgia",
    "Francesco": "Francesco Pellicani",
    "Andrea": "Andrea Fantuzzo",
    "Barbara": "Barbara Zuddas",
    "Alessandro": "Alessandro Demicheli",
    "Alessandrino": "Alessandro Demicheli",
    "Marco Benini": "Marco Benini",
}


# ──────────────────────────────────────────────────────────────────────
# PARSING DEI TEMPI
# ──────────────────────────────────────────────────────────────────────

def parse_time(raw) -> dict | None:
    """
    Converte un valore grezzo di tempo ed estrae note e distanze custom.
    Ritorna un dict: {'tempo': float, 'distanza': float (opzionale), 'nota': str} oppure None
    """
    if raw is None:
        return None

    # già numerico
    if isinstance(raw, (int, float)):
        if np.isnan(raw) if isinstance(raw, float) else False:
            return None
        return {'tempo': float(raw), 'distanza': None, 'nota': None}

    s = str(raw).strip()

    # valori nulli espliciti
    if s in ("", "/", "-", "❌", "x", "X"):
        return None

    # Estrazione distanza custom es. "50m:6.29" o "100: 12.70"
    custom_dist = None
    m_dist_explicit = re.match(r'^(\d+)\s*(?:m|mt)\s*[:\-]*\s*(.*)', s, re.IGNORECASE)
    m_dist_implicit = re.match(r'^(\d{2,4})\s*[:\-]+\s*(.*)', s, re.IGNORECASE)
    
    if m_dist_explicit:
        custom_dist = float(m_dist_explicit.group(1))
        s = m_dist_explicit.group(2)
    elif m_dist_implicit:
        custom_dist = float(m_dist_implicit.group(1))
        s = m_dist_implicit.group(2)

    # Estrazione stringhe di testo come Note (tutto ciò che non è numero o punteggiatura del tempo)
    # Rimuoviamo prima la punteggiatura tipica dei tempi per isolare vere "note" descrittive
    nota_str = re.sub(r'[\d\'\"\.,:]', '', s).strip()
    # Rimuoviamo le emoji e parentesi vuote rimaste
    nota_str = re.sub(r'[🦶🏻☠️❌✅💪🦁()]+', '', nota_str).strip()
    nota = nota_str if nota_str else None

    # Togliamo pure emoji e parentesi per estrarre il tempo pulito
    s_clean_time = s.split('/')[0].strip() # in caso di doppi tempi es: 15"40/12"55, prendo il primo
    s_clean_time = re.sub(r'[🦶🏻☠️❌✅💪🦁]+', '', s_clean_time)
    s_clean_time = s_clean_time.replace("''", '"') # Fix Marianna format
    s_clean_time = re.sub(r'\s*\(.*?\)\s*$', '', s_clean_time).strip()
    s_clean_time = re.sub(r'[a-zA-Z]+', '', s_clean_time).strip() # Togliamo le lettere

    if not s_clean_time or s_clean_time in ("/", "-"):
         return None

    tempo_float = None

    # Formato 1'04"  oppure  1'04"90
    m = re.match(r"(\d+)[''\u2032](\d+)[\"\u201d\u2033]?(\d+)?", s_clean_time)
    if m:
        minutes = int(m.group(1))
        secs = int(m.group(2))
        frac = int(m.group(3)) if m.group(3) else 0
        if frac > 0:
            frac_str = m.group(3)
            tempo_float = minutes * 60 + secs + frac / (10 ** len(frac_str))
        else:
            tempo_float = float(minutes * 60 + secs)

    if tempo_float is None:
        # Formato 7"12
        m = re.match(r'^(\d+)[\"\u201d\u2033](\d+)$', s_clean_time)
        if m:
            sec = int(m.group(1))
            frac_str = m.group(2)
            frac = int(frac_str)
            tempo_float = sec + frac / (10 ** len(frac_str))

    if tempo_float is None:
        # Formato 7"
        m = re.match(r'^(\d+)[\"\u201d\u2033]\s*$', s_clean_time)
        if m:
            tempo_float = float(m.group(1))

    if tempo_float is None:
        # Diretta 49.0
        s_clean = s_clean_time.replace(',', '.').replace('"', '').replace("'", '')
        try:
            tempo_float = float(s_clean)
        except ValueError:
            pass

    if tempo_float is None:
        nums = re.findall(r'\d+[\.,]?\d*', s_clean_time)
        if nums:
            try:
                tempo_float = float(nums[0].replace(',', '.'))
            except ValueError:
                pass

    if tempo_float is not None:
        return {'tempo': tempo_float, 'distanza': custom_dist, 'nota': nota}
    
    return None


# ──────────────────────────────────────────────────────────────────────
# PARSING DATA DA INTESTAZIONE TABELLA
# ──────────────────────────────────────────────────────────────────────

_MESI_IT = {
    'gennaio': 1, 'febbraio': 2, 'marzo': 3, 'aprile': 4,
    'maggio': 5, 'giugno': 6, 'luglio': 7, 'agosto': 8,
    'settembre': 9, 'ottobre': 10, 'novembre': 11, 'dicembre': 12,
}


def _parse_date_header(val) -> datetime | None:
    """Estrae una data dalla riga di intestazione."""
    if isinstance(val, datetime):
        return val

    if val is None:
        return None

    s = str(val).strip().lower()
    if not s:
        return None

    # Pattern tipo: "lunedi' 29 settembre 2025"
    m = re.search(r'(\d{1,2})\s+([a-zà-ú]+)\s+(\d{4})', s)
    if m:
        day = int(m.group(1))
        month_name = m.group(2)
        year = int(m.group(3))
        month = _MESI_IT.get(month_name)
        if month:
            try:
                return datetime(year, month, day)
            except ValueError:
                pass

    # Pattern dd/mm/yyyy
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    return None


# ──────────────────────────────────────────────────────────────────────
# CARICAMENTO DATI CORSA
# ──────────────────────────────────────────────────────────────────────

def load_running_data(filepath: str | Path) -> pd.DataFrame:
    """
    Legge 'Lavori Corsa.xlsx' e restituisce un DataFrame tidy:
      [Data, Atleta, Distanza, Tempo]

    Ogni foglio (mese) contiene più tabelle.
    Ogni tabella inizia con una riga data, poi un'intestazione con le distanze,
    e infine le righe degli atleti.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_date = None
        distances = []   # lista di (col_index, distanza_float)
        in_table = False

        for row_idx in range(1, ws.max_row + 1):
            row_vals = [cell.value for cell in ws[row_idx]]

            # Riga vuota → reset
            if all(v is None for v in row_vals):
                in_table = False
                distances = []
                continue

            first_val = row_vals[0]

            # Prova a interpretare come data
            date_parsed = _parse_date_header(first_val)
            if date_parsed is not None:
                current_date = date_parsed
                in_table = False
                distances = []
                continue

            # Riga di intestazione? (Prima cella = "Nome" o simile)
            if isinstance(first_val, str) and first_val.strip().lower() in (
                'nome', 'mezzofondisti', 'velocisti', 'quattrocentisti'
            ):
                # Se la riga contiene "Nome", la prossima colonna (o la stessa riga)
                # potrà avere le distanze
                if first_val.strip().lower() == 'nome':
                    distances = []
                    # Colonne distanza partono da indice 2 (dopo Nome, Cognome)
                    for ci in range(2, len(row_vals)):
                        dv = row_vals[ci]
                        if dv is not None:
                            try:
                                distances.append((ci, float(dv)))
                            except (ValueError, TypeError):
                                distances.append((ci, str(dv)))
                    in_table = True
                else:
                    # Riga di categoria ("MEZZOFONDISTI", ...), la prossima
                    # riga sarà l'header
                    in_table = False
                continue

            # Riga atleta (se siamo in una tabella)
            if in_table and distances and current_date is not None:
                nome = str(first_val).strip() if first_val else ""
                cognome = str(row_vals[1]).strip() if row_vals[1] else ""

                if not nome:
                    continue

                atleta = f"{nome} {cognome}".strip()

                for ci, dist in distances:
                    # Includi solo distanze numeriche
                    if not isinstance(dist, (int, float)):
                        continue
                    if ci < len(row_vals):
                        res = parse_time(row_vals[ci])
                        if res is not None:
                            actual_dist = res.get('distanza') if res.get('distanza') is not None else float(dist)
                            nota = res.get('nota')
                            records.append({
                                'Data': current_date,
                                'Atleta': atleta,
                                'Distanza': actual_dist,
                                'Tempo': res['tempo'],
                                'Note': nota if nota else ""
                            })

    wb.close()

    df = pd.DataFrame(records)
    if len(df) > 0:
        df['Data'] = pd.to_datetime(df['Data'])
        df['Tempo'] = df['Tempo'].astype(float)
    return df


# ──────────────────────────────────────────────────────────────────────
# CARICAMENTO VBT (XML DIRETTO)
# ──────────────────────────────────────────────────────────────────────

def load_vbt_data(filepath: str | Path) -> pd.DataFrame:
    """
    Legge 'VBT2026322.xlsx' direttamente dall'XML interno
    (aggira il crash openpyxl/calamine sui valori NaN).

    Restituisce un DataFrame con colonne tipizzate.
    Filtra le righe di test (atleta == 'user').
    """
    filepath = str(filepath)

    COLUMNS = [
        'Data',                      # A
        'Atleta',                    # B
        'Tempo_sessione',            # C
        'Esercizio',                 # D
        'Serie',                     # E
        'Ripetizioni',               # F
        'Carico',                    # G  (concentrico)
        'Vel_media',                 # H
        'Vel_max',                   # I
        'Tempo_acc_ms',              # J
        'Potenza_media',             # K
        'Potenza_max',               # L
        'Distanza_mm',               # M
        'Forza_max',                 # N
        '1RM',                       # O
        'Lavoro_J',                  # P
        'Eccentrico_carico',         # Q
        'Ecc_vel_media',             # R
        'Ecc_vel_max',               # S
        'Ecc_potenza_media',         # T
        'Ecc_potenza_max',           # U
        'Ecc_distanza_mm',           # V
        'Ecc_forza_max',             # W
        'Ecc_lavoro_J',              # X
    ]

    COL_LETTERS = [
        'A','B','C','D','E','F','G','H','I','J','K','L','M',
        'N','O','P','Q','R','S','T','U','V','W','X'
    ]
    col_map = {letter: idx for idx, letter in enumerate(COL_LETTERS)}

    with zipfile.ZipFile(filepath, 'r') as z:
        # Leggi shared strings
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            root = ET.parse(z.open('xl/sharedStrings.xml')).getroot()
            ns = re.match(r'\{.*\}', root.tag).group(0) if '{' in root.tag else ''
            for si in root.findall(f'{ns}si'):
                parts = []
                for t in si.iter(f'{ns}t'):
                    if t.text:
                        parts.append(t.text)
                shared_strings.append(''.join(parts))

        # Leggi il foglio
        root = ET.parse(z.open('xl/worksheets/sheet1.xml')).getroot()
        ns = re.match(r'\{.*\}', root.tag).group(0) if '{' in root.tag else ''

        rows_data = []
        first_row = True
        for row_elem in root.iter(f'{ns}row'):
            if first_row:
                first_row = False
                continue  # salta header

            row_dict = {c: None for c in COLUMNS}
            for cell in row_elem:
                ref = cell.get('r', '')
                col_letter = re.match(r'([A-Z]+)', ref)
                if not col_letter:
                    continue
                col_letter = col_letter.group(1)
                if col_letter not in col_map:
                    continue

                cell_type = cell.get('t', 'n')
                v_elem = cell.find(f'{ns}v')
                val = v_elem.text if v_elem is not None else None

                if cell_type == 's' and val is not None:
                    val = shared_strings[int(val)]

                col_idx = col_map[col_letter]
                row_dict[COLUMNS[col_idx]] = val

            rows_data.append(row_dict)

    df = pd.DataFrame(rows_data)

    # Filtra righe di test
    df = df[df['Atleta'].notna()]
    df = df[~df['Atleta'].str.strip().str.lower().isin(['user', ''])]

    # Forward-fill dei campi che si ripetono solo sulla prima riga di un set
    fill_cols = ['Data', 'Atleta', 'Tempo_sessione', 'Esercizio',
                 'Serie', 'Ripetizioni', 'Carico']
    for col in fill_cols:
        if col in df.columns:
            df[col] = df[col].replace('', np.nan).ffill()

    # Parsing del carico: "95KG" → 95.0
    def _parse_carico(v):
        if v is None or str(v).strip() == '':
            return np.nan
        s = str(v).upper().replace('KG', '').strip()
        try:
            return float(s)
        except ValueError:
            return np.nan

    for c in ['Carico', 'Eccentrico_carico', '1RM']:
        if c in df.columns:
            df[c] = df[c].apply(_parse_carico)

    # Converti colonne numeriche
    numeric_cols = [
        'Serie', 'Ripetizioni',
        'Vel_media', 'Vel_max', 'Tempo_acc_ms',
        'Potenza_media', 'Potenza_max', 'Distanza_mm',
        'Forza_max', 'Lavoro_J',
        'Ecc_vel_media', 'Ecc_vel_max',
        'Ecc_potenza_media', 'Ecc_potenza_max',
        'Ecc_distanza_mm', 'Ecc_forza_max', 'Ecc_lavoro_J',
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Parsing data
    if 'Data' in df.columns:
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce')

    # Pulizia esercizio
    if 'Esercizio' in df.columns:
        df['Esercizio'] = df['Esercizio'].str.strip()
        df.loc[df['Esercizio'] == 'Generale', 'Esercizio'] = 'General'

    df = df.reset_index(drop=True)
    return df


# ──────────────────────────────────────────────────────────────────────
# NORMALIZZAZIONE NOMI
# ──────────────────────────────────────────────────────────────────────

def normalize_names(df: pd.DataFrame, mapping: dict = None,
                    col: str = 'Atleta') -> pd.DataFrame:
    """
    Applica il dizionario di mappatura nomi.
    Se mapping è None usa il NOME_MAPPING di default.
    """
    if mapping is None:
        mapping = NOME_MAPPING

    df = df.copy()
    if col in df.columns:
        # Normalizza spazi
        df[col] = df[col].str.strip()
        # Title-case per uniformare (es. 'deidda' → 'Deidda')
        df[col] = df[col].str.title()
        # Costruisci mapping case-insensitive
        title_mapping = {k.strip().title(): v for k, v in mapping.items()}
        df[col] = df[col].replace(title_mapping)
    return df


# ──────────────────────────────────────────────────────────────────────
# FUNZIONE DI CARICAMENTO COMPLETA
# ──────────────────────────────────────────────────────────────────────

def load_all_data(running_path: str | Path, vbt_path: str | Path):
    """
    Carica e pulisce entrambi i dataset, normalizzando i nomi.
    Restituisce (df_running, df_vbt).
    """
    df_running = load_running_data(running_path)
    df_running = normalize_names(df_running)

    df_vbt = load_vbt_data(vbt_path)
    df_vbt = normalize_names(df_vbt)

    return df_running, df_vbt


# ──────────────────────────────────────────────────────────────────────
# TEST RAPIDO
# ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    base = Path(__file__).parent

    print("Caricamento dati corsa...")
    df_r = load_running_data(base / 'Lavori Corsa.xlsx')
    df_r = normalize_names(df_r)
    print(f"  Righe: {len(df_r)}")
    print(f"  Atleti: {sorted(df_r['Atleta'].unique())}")
    print(f"  Distanze: {sorted(df_r['Distanza'].unique())}")
    print(f"  Date: {df_r['Data'].min()} → {df_r['Data'].max()}")
    print(df_r.head(10))
    print()

    print("Caricamento dati VBT...")
    df_v = load_vbt_data(base / 'VBT2026322.xlsx')
    df_v = normalize_names(df_v)
    print(f"  Righe: {len(df_v)}")
    print(f"  Atleti: {sorted(df_v['Atleta'].unique())}")
    print(f"  Esercizi: {sorted(df_v['Esercizio'].dropna().unique())}")
    print(df_v.head(10))
